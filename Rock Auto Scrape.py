#! python3

#Requries PhantomJS
#Part numbers in file expected to be in excel file Column A, row 2+

from selenium import webdriver
from tkinter import *
import openpyxl, os, sys

file = filedialog.askopenfilename() #file browser window opens to select file
workbook = openpyxl.load_workbook(file)

if file == None:    #ends program if the user hits cancel
    exit()

try:
    workbook.save(file)  #checks to see if file can be saved
except PermissionError:
    messagebox.showinfo('Error','You must close the file')
    sys.exit()

sheet = workbook.active

browser = webdriver.PhantomJS()

sheet.cell(row=1, column = 2).value = 'Rebuild'
sheet.cell(row=1, column = 3).value = 'Core'

#Remanufactured auto parts are sold with a 'Core deposit'.  If the consumer returns
#with their original failed unit, the Core deposit is refunded.  That core is
#then a raw material for remanufacturers.

partCount = sheet.max_row - 1
i=2
while i <= partCount:
    part = str(sheet.cell(row=i,column=1).value)
    if part == 'None':
        i+=1
        partCount += 1
        continue
    browser.get('https://www.rockauto.com')
    searchElem = browser.find_element_by_xpath('//*[@id="searchinput[catalog]"]')
    searchElem.send_keys(part)
    searchElem.submit()
    try:
        priceElem = browser.find_element_by_xpath('//*[@id="dprice[3][td]"]/span')
        coreElem = browser.find_element_by_xpath('//*[@id="dcore[3][td]"]')
    except:
        i+=1
        continue
    sheet.cell(row=i, column =2).value = priceElem.text
    sheet.cell(row=i, column = 3).value = coreElem.text
    i+=1
    
browser.quit()

workbook.save(file)
messagebox.showinfo('Complete', 'Prices have been fetched and updated')

sys.exit()
